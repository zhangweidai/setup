from openpyxl import load_workbook
import download_financials
import os
import util
import fnmatch
import z
from collections import defaultdict, deque
from sortedcontainers import SortedSet
#want = ["Market Cap", "PE ratio", "Revenue per Share"]
want = ["Market Cap"]
icols = []
cdics = defaultdict(SortedSet)
bdics = SortedSet()

def proc(astock):
    global bdics
    path = download_financials.metrics(astock)
    if not os.path.exists(path):
        return
    wb = load_workbook(filename = path)
    ws = wb.active
    dates = list()
    for row in ws.values:
        title = None
        crow = list()
        save = list()
        for i, value in enumerate(row):
            if not dates:
                title = value
            if not title:
                title = value

            if value:
                crow.append(value)

            if title in want and type(value) != str :
                save.append(value)

        if not dates:
            dates = crow
            for i, date in enumerate(dates):
                icols.append(i+1)
#                if date.month == 12 or date.month == 11:
                
        if save:
            save.reverse()
            lists = []
            for i,ev in enumerate(save):
                try:
                    if ev == save[i-1]:
                        continue
                    change = round(ev/save[i-1],3)
                    lists.append(change)
                except:
                    continue

            leng = len(lists)
            if leng < 12:
                continue

            avg = sum(lists)/leng
            bdics.add((avg, astock))


def getXs():
    pattern = "*.xlsx"
    holds = []
    listOfFiles = os.listdir('../zen_dump/metrics')  
    for entry in listOfFiles:  
        if fnmatch.fnmatch(entry, pattern):
            holds.append(entry.split(".")[0])
    return holds

def proc2(astock, year = 2017):
    path = download_financials.metrics(astock)
    if not os.path.exists(path):
        return
    wb = load_workbook(filename = path)
    ws = wb.active
    dates = list()
    cols = list()
    for row in ws.values:
        title = None
        crow = list()
        save = list()
        dlist = list()
        for i, value in enumerate(row):
            if not dates:
                title = "Date"

            if not title:
                title = value

            if value:
                crow.append(value)

            if type(value) != str:
                if title in want and  i+1 in cols and value != 0.0:
                    save.append(value)
                elif title == "Date":
                    dlist.append(value)
                    try:
                        if value.year == year:
                            cols.append(i+1)
                    except:
                        pass
#                    try:
#                        print("value: {}".format( type(value.year)))
#                    except:
#                        pass

        if not dates:
            dates = crow
            for i, date in enumerate(dates):
                icols.append(i+1)
#                if date.month == 12 or date.month == 11:
                
#        if title == "Date":
#            print("dlist: {}".format( dlist))

        if save:
            avg = sum(save)/len(save)
            cdics[title].add((avg, astock))
#            print("title : {}".format( title ))
#            print("save: {}".format( sum(save)/len(save)))
def doit():
    global cdics
#    proc("MKTX")
    stocks = getXs()
    for astock in stocks:
        proc(astock)
    print("bdics: {}".format( bdics[-30:]))
    print("bdics: {}".format( bdics[:30]))
    z.setp( bdics[-30:], "changes")
    z.setp( bdics[:30], "changes2")
    return

    for i, item in enumerate(want):
        print("item : {}".format( item ))
        print("item : {}".format( cdics[item][-50:]))
        z.setp( cdics[item][-50:], "50_from_2017")

#    cdics = defaultdict(SortedSet)
#    for astock in stocks:
#        proc2(astock, 2012)
#
#    for i, item in enumerate(want):
#        print("item : {}".format( item ))
#        print("item : {}".format( cdics[item][-50:]))
#        z.setp( cdics[item][-50:], "50_from_2012")
        

        
if __name__ == '__main__':
    doit()
